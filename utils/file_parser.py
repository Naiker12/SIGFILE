import pandas as pd
import zipfile
import os
import tempfile
from datetime import datetime
import io
import numpy as np
import hashlib
import json

def parse_uploaded_file(filepath):
    """Parsea archivos subidos y retorna headers y datos"""
    if filepath.endswith(".xlsx"):
        return parse_excel_file(filepath)
    elif filepath.endswith(".zip"):
        return parse_zip_metadata(filepath)
    elif filepath.endswith(".ram"):
        return parse_ram_file(filepath)
    else:
        return [], []

def parse_excel_file(filepath):
    """Parsea archivos Excel manteniendo valores vacíos como None"""
    try:
        df = pd.read_excel(filepath, keep_default_na=True, na_values=[''])
        
        headers = [str(col) for col in df.columns]
        
        data = []
        for _, row in df.iterrows():
            row_dict = {}
            for i, header in enumerate(headers):
                value = row.iloc[i]
                if pd.isna(value):
                    row_dict[header] = None
                elif isinstance(value, (int, float)) and not pd.isna(value):
                    row_dict[header] = value
                else:
                    row_dict[header] = str(value).strip() if str(value).strip() != 'nan' else None
            
            # Agregar ID único para cada fila
            row_dict['_row_id'] = generate_row_id(row_dict)
            data.append(row_dict)
        
        return headers, data
        
    except Exception as e:
        print(f"Error leyendo archivo Excel: {e}")
        return [], []

def parse_zip_metadata(filepath):
    """Parsea metadatos de archivos ZIP incluyendo ZIPs anidados"""
    metadata = []

    try:
        with zipfile.ZipFile(filepath, 'r') as zf:
            for zinfo in zf.infolist():
                if zinfo.is_dir():
                    continue

                entry = {
                    "nombre": zinfo.filename,
                    "tamaño_original": format_size(zinfo.file_size),
                    "tamaño_comprimido": format_size(zinfo.compress_size),
                    "modificado": datetime(*zinfo.date_time).strftime("%Y-%m-%d %H:%M:%S"),
                    "tipo": get_file_type(zinfo.filename),
                    "ratio_compresion": calculate_compression_ratio(zinfo.file_size, zinfo.compress_size)
                }
                
                entry['_row_id'] = generate_row_id(entry)
                metadata.append(entry)

                if zinfo.filename.lower().endswith(".zip"):
                    try:
                        with zf.open(zinfo.filename) as inner_file:
                            nested_zip = zipfile.ZipFile(io.BytesIO(inner_file.read()))
                            for nested in nested_zip.infolist():
                                if nested.is_dir():
                                    continue
                                nested_entry = {
                                    "nombre": f"{zinfo.filename}/{nested.filename}",
                                    "tamaño_original": format_size(nested.file_size),
                                    "tamaño_comprimido": format_size(nested.compress_size),
                                    "modificado": datetime(*nested.date_time).strftime("%Y-%m-%d %H:%M:%S"),
                                    "tipo": f"ZIP anidado - {get_file_type(nested.filename)}",
                                    "ratio_compresion": calculate_compression_ratio(nested.file_size, nested.compress_size)
                                }
                                nested_entry['_row_id'] = generate_row_id(nested_entry)
                                metadata.append(nested_entry)
                            nested_zip.close()
                    except Exception as e:
                        print(f"Error leyendo ZIP anidado {zinfo.filename}: {e}")
                        continue

    except Exception as e:
        print(f"Error leyendo archivo ZIP: {e}")
        return [], []

    headers = ["nombre", "tamaño_original", "tamaño_comprimido", "modificado", "tipo", "ratio_compresion"]
    return headers, metadata

def parse_ram_file(filepath):
    """Parsea archivos RAM (texto delimitado) con detección automática de separadores"""
    try:
        encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
        lines = None
        
        for encoding in encodings:
            try:
                with open(filepath, "r", encoding=encoding, errors="ignore") as f:
                    lines = f.readlines()
                break
            except Exception:
                continue
        
        if not lines:
            return [], []

        separators = [",", ";", "\t", "|"]
        best_separator = ","
        max_columns = 0
        
        first_line = lines[0].strip()
        for sep in separators:
            if sep in first_line:
                column_count = len(first_line.split(sep))
                if column_count > max_columns:
                    max_columns = column_count
                    best_separator = sep

        headers = [header.strip().replace('"', '') for header in first_line.split(best_separator)]
        
        data = []
        for line_num, line in enumerate(lines[1:], 2):
            line = line.strip()
            if not line:  
                continue
                
            parts = [part.strip().replace('"', '') for part in line.split(best_separator)]
            
            while len(parts) < len(headers):
                parts.append(None)
            parts = parts[:len(headers)]
            
            row_dict = {}
            for i, header in enumerate(headers):
                value = parts[i] if i < len(parts) else None

                if value is None or value == '' or value.lower() in ['null', 'nan', 'none']:
                    row_dict[header] = None
                else:
                    row_dict[header] = value
            
            # Agregar ID único
            row_dict['_row_id'] = generate_row_id(row_dict)
            data.append(row_dict)

        return headers, data

    except Exception as e:
        print(f"Error leyendo archivo RAM: {e}")
        return [], []

def generate_row_id(row_dict):
    """Genera un ID único para una fila basado en su contenido"""
    content = json.dumps({k: v for k, v in row_dict.items() if k != '_row_id'}, 
                        sort_keys=True, default=str)
    return hashlib.md5(content.encode()).hexdigest()[:12]

def remove_duplicates(data, key_columns=None):
    """Elimina filas duplicadas basándose en columnas específicas o en todo el contenido"""
    if not data:
        return data
    
    seen = set()
    unique_data = []
    
    for row in data:
        if key_columns:
            key_values = tuple(str(row.get(col, '')) for col in key_columns)
        else:
            key_values = row.get('_row_id', generate_row_id(row))
        
        if key_values not in seen:
            seen.add(key_values)
            unique_data.append(row)
    
    return unique_data

def apply_filters(data, filters):
    """Aplica filtros a los datos"""
    if not data or not filters:
        return data
    
    filtered_data = []
    for row in data:
        include_row = True
        
        for column, filter_value in filters.items():
            if column not in row:
                continue
                
            row_value = str(row[column]).lower() if row[column] is not None else ''
            filter_value = str(filter_value).lower()
            
            if filter_value and filter_value not in row_value:
                include_row = False
                break
        
        if include_row:
            filtered_data.append(row)
    
    return filtered_data

def save_processed_data(data, filepath):
    """Guarda los datos procesados en un archivo temporal para persistencia"""
    try:
        temp_file = tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.json')
        json.dump(data, temp_file, default=str, indent=2)
        temp_file.close()
        return temp_file.name
    except Exception as e:
        print(f"Error guardando datos procesados: {e}")
        return None

def load_processed_data(filepath):
    """Carga datos procesados desde un archivo temporal"""
    try:
        with open(filepath, 'r') as f:
            return json.load(f)
    except Exception as e:
        print(f"Error cargando datos procesados: {e}")
        return []

def format_size(size_bytes):
    """Convierte bytes a formato legible"""
    if size_bytes == 0:
        return "0 B"
    
    size_names = ["B", "KB", "MB", "GB", "TB"]
    i = 0
    while size_bytes >= 1024 and i < len(size_names) - 1:
        size_bytes /= 1024.0
        i += 1
    
    return f"{size_bytes:.1f} {size_names[i]}"

def get_file_type(filename):
    """Determina el tipo de archivo basado en la extensión"""
    extension = filename.lower().split('.')[-1] if '.' in filename else ''
    
    file_types = {
        'txt': 'Texto',
        'csv': 'CSV',
        'xlsx': 'Excel',
        'xls': 'Excel',
        'pdf': 'PDF',
        'doc': 'Word',
        'docx': 'Word',
        'ppt': 'PowerPoint',
        'pptx': 'PowerPoint',
        'jpg': 'Imagen',
        'jpeg': 'Imagen',
        'png': 'Imagen',
        'gif': 'Imagen',
        'zip': 'ZIP',
        'rar': 'RAR',
        '7z': '7-Zip',
        'mp3': 'Audio',
        'mp4': 'Video',
        'avi': 'Video',
        'mov': 'Video'
    }
    
    return file_types.get(extension, 'Archivo')

def calculate_compression_ratio(original_size, compressed_size):
    """Calcula el ratio de compresión"""
    if original_size == 0:
        return "0%"
    
    ratio = ((original_size - compressed_size) / original_size) * 100
    return f"{ratio:.1f}%"

def export_to_excel(data, selected_columns=None, remove_duplicates_flag=True, filters=None):
    """Exporta datos a Excel con columnas seleccionadas, eliminación de duplicados y filtros aplicados"""
    try:
        if not data:
            return None
        if filters:
            data = apply_filters(data, filters)
        if remove_duplicates_flag:
            data = remove_duplicates(data)
    
        if selected_columns:
            filtered_data = []
            for row in data:
                filtered_row = {col: row.get(col) for col in selected_columns if col != '_row_id'}
                filtered_data.append(filtered_row)
            data = filtered_data
        else:
            # Remover _row_id de la exportación
            data = [{k: v for k, v in row.items() if k != '_row_id'} for row in data]
        
        if not data:
            print("No hay datos para exportar después de aplicar filtros")
            return None
        
        df = pd.DataFrame(data)
        df = df.fillna('')
        
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        
        with pd.ExcelWriter(temp_file.name, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Datos', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Datos']
            
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="007A33", end_color="007A33", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                
                column_letter = cell.column_letter
                max_length = max(len(str(column_title)), 15)
                worksheet.column_dimensions[column_letter].width = min(max_length, 50)
        
        return temp_file.name
        
    except Exception as e:
        print(f"Error exportando a Excel: {e}")
        return None

def get_data_summary(data):
    """Genera un resumen de los datos"""
    if not data:
        return {}
    
    summary = {
        'total_rows': len(data),
        'total_columns': len(data[0].keys()) if data else 0,
        'duplicates': len(data) - len(remove_duplicates(data)),
        'empty_values': 0
    }
    
    # Contar valores vacíos
    for row in data:
        for value in row.values():
            if value is None or value == '':
                summary['empty_values'] += 1
    
    return summary